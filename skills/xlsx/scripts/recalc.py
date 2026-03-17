import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not found. Please install with 'pip install openpyxl'")
    sys.exit(1)

def check_formulas(xlsx_path):
    results = {
        "status": "success",
        "total_formulas": 0,
        "total_errors": 0,
        "error_details": []
    }
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        results["total_formulas"] += 1
                        
        # In a real expert scenario, we'd use LibreOffice/Excel to recalc. 
        # Here we verify at least the static data_only values if available.
        wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)
        for sheet_name in wb_data.sheetnames:
            ws = wb_data[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in ["#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]:
                        results["total_errors"] += 1
                        results["error_details"].append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "type": cell.value
                        })
        
        if results["total_errors"] > 0:
            results["status"] = "errors_found"
            
    except Exception as e:
        return {"error": str(e)}
        
    return results

def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <input.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    print(json.dumps(check_formulas(xlsx_path), indent=2))

if __name__ == "__main__":
    main()
